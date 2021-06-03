import pandas as pd
import numpy as np
import openpyxl
import random


def normalize(data):
    """
    归一化
    :param data: 输入数据
    :return:
    """
    data_max, data_min = data.max(axis=0), data.min(axis=0)
    data = (data - data_min) / (data_max - data_min)
    return data


def binary_classification(data):
    """
    二分类（normal与其他类型）
    :param data: 原始数据
    :return:
    """
    # 将标签改为 0 与 1
    # 时间与id之外的其余数据归一化
    attributes = [col for col in data]
    pre_data = {}
    for attribute in attributes:
        tmp = data[attribute].values
        if attribute == ' who CH' or attribute == ' id' or attribute == ' Time':
            tmp = tmp
        elif isinstance(tmp[0], float) or isinstance(tmp[0], np.int64):
            # print(attribute, tmp[:5])
            tmp = normalize(tmp)
            # print(attribute, tmp[:5])
        elif attribute == 'Attack type':
            label = [1 if j == 'Normal' else 0 for j in tmp]
            tmp = np.array(label)
        else:
            raise IndexError
        pre_data[attribute] = tmp
    # 随机数生成
    sample_train_index = sorted(random.sample(range(len(pre_data['Attack type'])),
                                              int(0.1 * len(pre_data['Attack type']))))
    sample_test_index = sorted(random.sample(range(len(pre_data['Attack type'])),
                                             int(0.1 * len(pre_data['Attack type']))))
    # sample_test_index = sorted(list(set(range(len(pre_data['Attack type']))) -
    #                                 set(sample_train_index)))

    # 数据集分割
    train_data = {}
    test_data = {}
    for key, value in pre_data.items():
        tmp_train = [value[i] for i in sample_train_index]
        tmp_test = [value[i] for i in sample_test_index]
        train_data[key] = tmp_train
        test_data[key] = tmp_test

    # 保存为 xls 文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet('train data', index=0)
    for i in range(len(attributes)):  # 写入列名
        outws.cell(1, i + 1).value = attributes[i]
    outwb.save("binary_wsn-ds.xlsx")
    col = 1  # 按列写入内容
    for key, data in train_data.items():
        for row in range(0, len(data)):
            outws.cell(row + 2, col).value = data[row]  # 写文件
        col += 1
        outwb.save("binary_wsn-ds.xlsx")  # 一定要记得保存

    outws = outwb.create_sheet('test data', index=1)
    for i in range(len(attributes)):  # 写入列名
        outws.cell(1, i + 1).value = attributes[i]
    outwb.save("binary_wsn-ds.xlsx")
    col = 1  # 按列写入内容
    for key, data in test_data.items():
        for row in range(0, len(data)):
            outws.cell(row + 2, col).value = data[row]  # 写文件
        col += 1
        outwb.save("binary_wsn-ds.xlsx")  # 一定要记得保存


def random_sampling(data):
    """
    随机选取，保存为xls（0.7：0.3），按标签存储，按顺序保留所有属性
    :param data: 原始数据
    :return:
    """
    attributes = [col for col in data]
    from sklearn.model_selection import train_test_split
    normal = data[data['Attack type'] == 'Normal']
    blackhole = data[data['Attack type'] == 'Blackhole']
    grayhole = data[data['Attack type'] == 'Grayhole']
    flooding = data[data['Attack type'] == 'Flooding']
    tdma = data[data['Attack type'] == 'TDMA']

    # 分割
    normal_train, normal_test = train_test_split(normal, test_size=0.3)
    normal_train, normal_test = normal_train.sort_index(ascending=True), normal_test.sort_index(ascending=True)

    blackhole_train, blackhole_test = train_test_split(blackhole, test_size=0.3)
    blackhole_train, blackhole_test = blackhole_train.sort_index(ascending=True), blackhole_test.sort_index(
        ascending=True)

    grayhole_train, grayhole_test = train_test_split(grayhole, test_size=0.3)
    grayhole_train, grayhole_test = grayhole_train.sort_index(ascending=True), grayhole_test.sort_index(ascending=True)

    flooding_train, flooding_test = train_test_split(flooding, test_size=0.3)
    flooding_train, flooding_test = flooding_train.sort_index(ascending=True), flooding_test.sort_index(ascending=True)

    tdma_train, tdma_test = train_test_split(tdma, test_size=0.3)
    tdma_train, tdma_test = tdma_train.sort_index(ascending=True), tdma_test.sort_index(ascending=True)

    dataset = {"blackhole_train": blackhole_train, "blackhole_test": blackhole_test,
               "grayhole_train": grayhole_train, "grayhole_test": grayhole_test,
               "flooding_train": flooding_train, "flooding_test": flooding_test,
               "tdma_train": tdma_train, "tdma_test": tdma_test,
               "normal_train": normal_train, "normal_test": normal_test, }
    # 保存为 xls 文件
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    index = 0
    for key, value in dataset.items():
        outws = outwb.create_sheet(key, index=index)  # 在将写的文件创建sheet
        for row in range(1, len(value) + 1):
            for col in range(1, len(attributes) + 1):
                outws.cell(row, col).value = value.iloc[row - 1][col - 1]  # 写文件
        index += 1
        outwb.save("random_wsn-ds.xls")  # 一定要记得保存


def run():
    data = pd.read_csv("WSN-DS.csv")
    attributes = [column for column in data]
    print(attributes)
    # random_sampling(data)
    binary_classification(data)


if __name__ == '__main__':
    run()

# 特征说明
# 1 Node ID             节点ID号
# 2 Time                当前节点的运行时间
# 3 IS CH               用于标志节点是否为簇头CH，是为1，普通节点为0
# 4 Who CH              簇头CH的ID
# 5 Distance to CH      当前节点与簇头CH的距离
# 6 Energy consumption  消耗的能量
# 7 ADV CH send         簇头CH发送到节点的消息
# 8 ADV CH receives     簇头CH广播的消息数量
# 9 Join REQ send       节点发送给CH连接请求消息的数量
# 10 Join REQ receive   CH从节点接收到的连接请求消息数
# 11 ADV SCH send       TDMA调度发送到节点的广播消息数
# 12 ADV SCH receives   CH接收到的TDMA调度消息数量
# 13 Rank               节点TDMA调度的顺序
# 14 Data sent          从普通节点发送到其CH的数据包的数量
# 15 Data received      节点从CH接收的数据包的数量
# 16 Data sent to BS    发送到BS的数据包的数量
# 17 Distance CH to BS  CH和BS之间的距离
# 18 Send Code          簇头发送的代码消息
# 19 Attack Type        分类标记
